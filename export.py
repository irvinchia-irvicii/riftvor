"""export.py — xlsx export of a comparison table + purchase plans
(openpyxl, export shape ported from 3vor Fetch)."""
from __future__ import annotations

import io
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import basket
from config import STORE_ORDER

_HEADER_FILL = PatternFill("solid", fgColor="1F2430")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BEST_FILL = PatternFill("solid", fgColor="C6EFCE")
_OOS_FONT = Font(color="999999", italic=True)
_SECTION_FONT = Font(bold=True, size=12)
_STORE_FONT = Font(bold=True)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width


def _comparison_sheet(wb: Workbook, result: dict, order: list[str],
                      names: dict[str, str]) -> None:
    ws = wb.active
    ws.title = "Comparison"
    ws.append(["Card", "Set", "No.", "Finish", "Qty", "Best (SGD)"]
              + [names.get(k, k) for k in order])
    _style_header(ws)
    for row in result.get("rows", []):
        values = [row["name"], row["set_code"], row["number"], row["finish"],
                  row["qty"], row["best_price"]]
        for key in order:
            cell = row["stores"].get(key)
            values.append(cell["price"] if cell else None)
        ws.append(values)
        excel_row = ws.max_row
        for i, key in enumerate(order):
            cell_data = row["stores"].get(key)
            cell = ws.cell(row=excel_row, column=7 + i)
            if cell_data:
                cell.hyperlink = cell_data["url"]
                cell.number_format = "0.00"
                if not cell_data["in_stock"]:
                    cell.font = _OOS_FONT
                elif (row["best_price"] is not None
                      and cell_data["price"] == row["best_price"]):
                    cell.fill = _BEST_FILL
        ws.cell(row=excel_row, column=6).number_format = "0.00"
    _widths(ws, [32, 6, 8, 8, 5, 10] + [14] * len(order))
    ws.freeze_panes = "A2"


def _plan_sheet(wb: Workbook, plans: dict) -> None:
    """Summary of every plan, then the cheapest basket store by store —
    the sheet you actually shop from."""
    all_plans = list(plans.get("plans", []))
    singles = list(plans.get("single_store", []))[:3]
    if not all_plans:
        return
    names = plans.get("store_names", {})
    ws = wb.create_sheet("Plan")
    ws.append(["Plan", "Total (SGD)", "Stores", "Lines filled",
               "Cards", "vs cheapest"])
    _style_header(ws)
    for plan in all_plans + singles:
        ws.append([plan["label"], plan["total"], plan["store_count"],
                   f"{plan['lines_filled']}/{plan['lines_total']}",
                   plan["cards_total"],
                   plan.get("premium") if plan.get("premium") else None])
        ws.cell(row=ws.max_row, column=2).number_format = "0.00"
        ws.cell(row=ws.max_row, column=6).number_format = "0.00"

    cheapest = all_plans[0]
    ws.append([])
    ws.append([f"{cheapest['label']} — what to buy"])
    ws.cell(row=ws.max_row, column=1).font = _SECTION_FONT
    ws.append(["Store", "Card", "Card no.", "Finish", "Qty", "Unit", "Line"])
    for cell in ws[ws.max_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for store in cheapest["by_store"]:
        first = True
        for line in store["lines"]:
            ws.append([names.get(store["store"], store["store"]) if first else "",
                       line["name"], line["card_key"], line["finish"],
                       line["qty"], line["unit_price"], line["line_cost"]])
            excel_row = ws.max_row
            if first:
                ws.cell(row=excel_row, column=1).font = _STORE_FONT
                first = False
            ws.cell(row=excel_row, column=2).hyperlink = line["url"]
            ws.cell(row=excel_row, column=6).number_format = "0.00"
            ws.cell(row=excel_row, column=7).number_format = "0.00"
        ws.append(["", "", "", "", "", "Subtotal", store["subtotal"]])
        ws.cell(row=ws.max_row, column=6).font = _STORE_FONT
        ws.cell(row=ws.max_row, column=7).font = _STORE_FONT
        ws.cell(row=ws.max_row, column=7).number_format = "0.00"

    ws.append(["", "", "", "", "", "TOTAL", cheapest["total"]])
    ws.cell(row=ws.max_row, column=6).font = _SECTION_FONT
    ws.cell(row=ws.max_row, column=7).font = _SECTION_FONT
    ws.cell(row=ws.max_row, column=7).number_format = "0.00"

    if cheapest.get("missing"):
        ws.append([])
        ws.append(["Nothing in stock anywhere:"])
        ws.cell(row=ws.max_row, column=1).font = _SECTION_FONT
        for raw in cheapest["missing"]:
            ws.append([raw])
    if plans.get("assumes_stock_depth"):
        ws.append([])
        ws.append(["Note: stores publish in-stock, not how many — multi-copy "
                   "lines assume the full quantity is available."])
    _widths(ws, [18, 32, 12, 9, 6, 10, 10])


def comparison_xlsx(result: dict) -> io.BytesIO:
    order = result.get("store_order") or STORE_ORDER
    names = basket.store_names()
    wb = Workbook()
    _comparison_sheet(wb, result, order, names)
    _plan_sheet(wb, result.get("basket") or {})
    if result.get("unmatched"):
        ws = wb.create_sheet("Unmatched")
        ws.append(["Query"])
        _style_header(ws)
        for entry in result["unmatched"]:
            ws.append([entry["raw"]])
        _widths(ws, [40])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def filename() -> str:
    return f"riftvor_comparison_{time.strftime('%Y%m%d_%H%M')}.xlsx"
